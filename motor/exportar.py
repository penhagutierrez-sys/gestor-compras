"""
Exporta las Órdenes de Compra sugeridas a un Excel ordenado y legible.
"""
from datetime import datetime

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import config

# Columnas que mostramos en el Excel y su título "bonito".
COLUMNAS = {
    "CODIGO": "Código",
    "PRODUCTO": "Producto",
    "RUBRO": "Rubro",
    "MARCA": "Marca",
    "ABC": "ABC",
    "XYZ": "XYZ",
    "PRONOSTICO_MENSUAL": "Pronóstico/mes",
    "STOCK_ACTUAL": "Stock actual",
    "SUGERIDO_PEDIR": "Sugerido pedir",
    "COSTO_UNIT": "Costo unit.",
    "MONTO_ESTIMADO": "Monto estimado",
}


def exportar_excel(ordenes, ruta=None):
    """Escribe el Excel y devuelve la ruta del archivo creado."""
    config.CARPETA_SALIDAS.mkdir(exist_ok=True)
    if ruta is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        ruta = config.CARPETA_SALIDAS / f"OC_sugeridas_{ts}.xlsx"

    tabla = ordenes[list(COLUMNAS.keys())].copy()
    tabla["PRONOSTICO_MENSUAL"] = tabla["PRONOSTICO_MENSUAL"].round(1)
    tabla = tabla.rename(columns=COLUMNAS)

    tabla.to_excel(ruta, index=False, sheet_name="OC sugeridas")
    _dar_formato(ruta, tabla)
    return ruta


def _dar_formato(ruta, tabla):
    """Encabezado en negrita, filtros, panel congelado y anchos de columna."""
    import openpyxl

    wb = openpyxl.load_workbook(ruta)
    ws = wb.active

    azul = PatternFill("solid", fgColor="1F4E78")
    for col, titulo in enumerate(tabla.columns, start=1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = azul
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"  # deja el encabezado fijo al hacer scroll
    ws.auto_filter.ref = ws.dimensions  # filtros en cada columna

    anchos = {"Producto": 42, "Rubro": 16, "Marca": 16, "Código": 16}
    for col, titulo in enumerate(tabla.columns, start=1):
        letra = get_column_letter(col)
        ws.column_dimensions[letra].width = anchos.get(titulo, 14)

    # Miles para montos y stock.
    for fila in ws.iter_rows(min_row=2):
        for celda in fila:
            titulo = tabla.columns[celda.column - 1]
            if titulo in ("Monto estimado", "Costo unit.", "Stock actual", "Sugerido pedir"):
                celda.number_format = "#,##0"

    wb.save(ruta)
